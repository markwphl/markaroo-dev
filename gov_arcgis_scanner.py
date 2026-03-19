#!/usr/bin/env python3
"""
Government ArcGIS Feature Layer Scanner
========================================
Discovers and catalogs GIS feature layers from local government ArcGIS REST
Services directories, filters for planning/development-related layers,
deduplicates, and exports results to Excel and Markdown.

Designed to run from a private corporate intranet in production.

## Architecture — Two-Tier ArcGIS Endpoint Discovery
------------------------------------------------------
When running in "discovery" mode (Path B), the scanner uses a two-tier
strategy to locate the jurisdiction's ArcGIS REST Services Directory URL:

  Tier 1 — Fast-Path Probe (FREE, no API cost)
      Tests common subdomain/path patterns (gis.<domain>, maps.<domain>,
      etc.) with lightweight HTTP HEAD/GET requests. If any return a valid
      ArcGIS services page, discovery is complete.

  Tier 2 — LLM Web Search (requires ANTHROPIC_API_KEY)
      If Tier 1 fails, calls the Claude API with the ``web_search`` server
      tool.  Claude searches the public web using the jurisdiction name and
      URL patterns from ``docs/planning-layer-pattern-skill-v2.md`` to find
      either the REST Services Directory URL or individual feature layer
      URLs (from which the directory root is derived).

## Configurable Variables (update as needed)
---------------------------------------------
The following module-level constants can be changed by your deployment team:

  REQUEST_TIMEOUT   — HTTP timeout per request (seconds).  Default: 20
  USER_AGENT        — Browser User-Agent header sent with every request.
  ARCGIS_REST_PATTERNS — Regex list used to recognise ArcGIS REST URLs.
  _LLM_MODEL        — Claude model ID used for web search (env var
                      ``ARCGIS_SCANNER_MODEL``).  Default: claude-sonnet-4-6

## Environment Variables
-------------------------
  ANTHROPIC_API_KEY       — (Required for Tier 2) Claude API key.
                            Obtain from https://console.anthropic.com
  ARCGIS_SCANNER_MODEL    — (Optional) Override the Claude model used for
                            web search.  Default: claude-sonnet-4-6

## Security Notes
------------------
  • The API key is read from the environment and NEVER logged, printed, or
    included in output files.
  • All URLs received from external sources (LLM responses, crawled pages)
    are validated against an allowlist of schemes (http/https only) and
    checked for private/reserved IP ranges (SSRF protection).
  • User-supplied jurisdiction names are sanitised to alphanumeric + spaces
    before being sent to the LLM.
  • The web application (web_app.py) validates job IDs and prevents path
    traversal in file download endpoints.
"""

import argparse
import ipaddress
import json
import logging
import os
import re
import socket
import sys
import time
import threading
from collections import defaultdict
from difflib import SequenceMatcher
from typing import Optional
from urllib.parse import urljoin, urlparse

import requests
import urllib3

# Suppress InsecureRequestWarning for government sites with bad SSL certs.
# Many government servers have misconfigured or self-signed certificates;
# this prevents noisy warnings in production logs while still allowing
# SSL-fallback requests where needed.
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

# Module logger — does NOT log sensitive data (API keys, credentials).
logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Security — SSRF Protection
# ---------------------------------------------------------------------------
# Prevent Server-Side Request Forgery (SSRF) by blocking requests to private,
# loopback, and reserved IP ranges.  All URLs obtained from untrusted sources
# (LLM responses, crawled HTML) MUST pass ``is_safe_url()`` before being
# fetched.

# Maximum allowed length for any URL we process (prevents memory abuse)
_MAX_URL_LENGTH = 2048

# Maximum allowed length for jurisdiction name input
_MAX_JURISDICTION_NAME_LENGTH = 200


def _resolve_hostname(hostname: str) -> list[str]:
    """Resolve a hostname to its IP addresses.  Returns an empty list on
    failure rather than raising, so callers can treat unresolvable hosts
    as unsafe."""
    try:
        return [info[4][0] for info in socket.getaddrinfo(hostname, None)]
    except (socket.gaierror, OSError):
        return []


def is_safe_url(url: str) -> bool:
    """Return True if *url* is safe to fetch (public internet, http/https).

    Rejects:
      • Non-http(s) schemes (file://, ftp://, gopher://, etc.)
      • URLs longer than ``_MAX_URL_LENGTH``
      • Hostnames that resolve to private/reserved/loopback IP ranges
      • Bare IP addresses in private ranges
    """
    if not url or len(url) > _MAX_URL_LENGTH:
        return False

    try:
        parsed = urlparse(url)
    except ValueError:
        return False

    # Only allow http and https schemes
    if parsed.scheme not in ("http", "https"):
        return False

    hostname = parsed.hostname
    if not hostname:
        return False

    # Check if the hostname is a raw IP address first
    try:
        addr = ipaddress.ip_address(hostname)
        if addr.is_private or addr.is_loopback or addr.is_reserved or addr.is_link_local:
            return False
    except ValueError:
        # Not a raw IP — resolve DNS and check each resulting address
        resolved = _resolve_hostname(hostname)
        for ip_str in resolved:
            try:
                addr = ipaddress.ip_address(ip_str)
                if addr.is_private or addr.is_loopback or addr.is_reserved or addr.is_link_local:
                    return False
            except ValueError:
                continue

    return True


def sanitize_jurisdiction_name(name: str) -> str:
    """Sanitise a jurisdiction name to prevent prompt injection or unexpected
    content being forwarded to the LLM.

    Allows only letters, digits, spaces, hyphens, periods, and apostrophes.
    Truncates to ``_MAX_JURISDICTION_NAME_LENGTH`` characters.
    """
    if not name:
        return ""
    # Strip characters that aren't alphanumeric, space, hyphen, period, or apostrophe
    cleaned = re.sub(r"[^a-zA-Z0-9 .'\-]", "", name)
    # Collapse multiple spaces
    cleaned = re.sub(r"\s+", " ", cleaned).strip()
    return cleaned[:_MAX_JURISDICTION_NAME_LENGTH]


# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
# These values can be adjusted by your deployment team to tune performance
# and behaviour for your network environment.

REQUEST_TIMEOUT = 20  # seconds — HTTP timeout per request
# User-Agent header — UPDATE this to match your organisation's policy.
# Some government servers block requests without a recognised browser UA.
USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
)
HEADERS = {"User-Agent": USER_AGENT}

# ArcGIS REST Services Directory URL patterns — regexes used to recognise
# valid ArcGIS REST endpoint URLs in crawled HTML and LLM responses.
ARCGIS_REST_PATTERNS = [
    re.compile(r"https?://[^\s\"'<>]+/arcgis/rest/services", re.IGNORECASE),
    re.compile(r"https?://services\d*\.arcgis\.com/[A-Za-z0-9]+/ArcGIS/rest/services", re.IGNORECASE),
    re.compile(r"https?://[^\s\"'<>]+/rest/services(?:/|$)", re.IGNORECASE),
]

# ---------------------------------------------------------------------------
# Tier 1 — Service Path Signals (per planning-layer-pattern-skill-v2.md)
# ---------------------------------------------------------------------------

# Department identifiers in service path (+5 confidence)
TIER1_DEPARTMENT_TOKENS = [
    "comdev", "cdd", "planning", "landdev", "communitydevelopment",
    "community_development", "energov", "plu", "rois", "p_d", "cd",
]

# Land use / zoning service name tokens (+4 confidence)
TIER1_LANDUSE_TOKENS = [
    "zoning", "land_use", "landuse", "future_land_use", "general_plan",
    "residential_zoning", "masterplan", "master_plan", "master_planned",
    "historic_district", "nrhp", "subdivision", "tif_parcels", "tif_zones",
    "housing_element", "housingelement", "specific_plans", "specificplans",
    "precise_plan", "preciseplan", "opportunity_sites", "urban_growth",
    "ugb", "parcels", "form_based_code", "form.based.code",
    "comprehensive_plan", "flum", "special_area_plan", "mello_roos",
    "melloroos", "mello-roos", "zcu",
]

# Development / planning service name tokens (+3 confidence)
TIER1_DEVELOPMENT_TOKENS = [
    "development", "agreement", "growth", "plan", "development_agreement",
    "outside_development", "downtown", "dtmasterplan", "envision",
    "growth_framework", "corridor_plan", "sector_plan",
]

# Combined list for simple boolean check (backward compat)
SERVICE_KEYWORDS = (
    TIER1_DEPARTMENT_TOKENS + TIER1_LANDUSE_TOKENS + TIER1_DEVELOPMENT_TOKENS
)

# ---------------------------------------------------------------------------
# Tier 2 — Layer Name Keyword Signals (10 semantic clusters)
# ---------------------------------------------------------------------------

# Cluster A: Zoning and Land Use Regulation
CLUSTER_A = [
    "zoning", "zone", "zone overlay", "zone map", "overlay", "land use",
    "future land use", "general plan", "comprehensive plan", "flum",
    "planned development", "form-based code", "form based code", "infill",
    "residential zoning", "non attainment", "ugb", "urban growth boundary",
    "zoning code update", "zcu", "hillside code", "hillside regulation",
    "heights", "tree zone", "tree preservation",
    # California SB 9 / AB 2923
    "sb 9", "ab 2923", "two-unit overlay", "two unit overlay",
    "lot split zone", "lot split", "state density bonus", "density bonus",
    "ministerial zone", "ministerial",
    # ADU
    "adu overlay", "adu zone", "accessory dwelling", "junior adu", "adu",
    # Southeast
    "udo", "unified development ordinance", "tnd",
    "traditional neighborhood", "rural preservation",
    # Midwest
    "overlay district", "conservation district",
]

# Cluster B: Development Review and Entitlements
CLUSTER_B = [
    "subdivision", "subdivisions", "subdivision sections",
    "special use permit", "development agreement", "agreement area",
    "pace agreement", "tdr", "transfer of development rights",
    "tif", "tax increment", "annexation", "annex", "deannex",
    "outside development", "planned development", "planned unit development",
    "pud", "housing element", "housing element sites", "by-right",
    "opportunity sites", "project development", "project developments",
    "project pipeline", "innovation parcel",
    # Housing
    "bmr", "below market rate", "low income housing", "affordable housing",
    "cfd", "cfd parcels", "mello-roos", "mello_roos", "melloroos",
    # Southeast / Regional
    "vad", "voluntary agricultural district", "voluntary agricultural",
    # California RHNA
    "rhna", "rhna sites", "rhna allocation", "rezoning rhna",
    # Midwest
    "enterprise zone", "special assessment district", "tax abatement",
    # Texas
    "pdd", "planned development district", "sup", "specific use permit",
    "reinvestment zone", "tirz", "4b sales tax",
    # Southwest
    "pad", "planned area development",
    # General
    "conservation subdivision", "activity center",
    "plat", "entitlement", "setback", "permits",
]

# Cluster C: Comprehensive / Master Planning
CLUSTER_C = [
    "master plan", "master planned", "general plan", "comprehensive plan",
    "future land use", "growth framework", "character areas",
    "special area plan", "specific plans", "specific plan", "precise plan",
    "precise plan area", "civic master plan", "downtown plan",
    "town center plan", "viewplane", "mtn viewplane", "greenline",
    "sphere of influence", "urban growth boundary",
    "transit priority area", "high quality transit corridor", "hqtc",
    "transit buffer", "midtown boundary", "corridor plan", "corridor",
    "small area plan", "area plan", "sector plan",
    # Southwest
    "rural planning area", "thoroughfare plan",
]

# Cluster D: Historic Preservation
CLUSTER_D = [
    "historic district", "historic sites", "historic buildings",
    "historic inventory", "historical inventory", "historical district",
    "nrhp", "character areas", "landmark", "local historic district",
    "national historic district", "heritage district", "preservation district",
]

# Cluster E: Environmental Overlay (Planning-Regulated)
CLUSTER_E = [
    "stream margin", "wildfire hazard", "wildfire district", "wildfire zone",
    "wildfire buffer", "fire zone", "fire buffer",
    "wui", "wildland-urban interface",
    "esa", "environmentally sensitive", "hallam bluff", "non attainment",
    "wetlands", "wetland buffer", "riparian buffer", "riparian zone",
    "riparian corridor", "watershed",
    "flood zone", "floodplain", "floodway", "flood plain",
    "fema flood", "fema floodplain",
    "100 year flood", "100-year flood", "500 year flood", "500-year flood",
    "100 year floodway", "special flood hazard area", "sfha",
    "flood zone parcels",
    "tree preservation", "tree canopy", "agricultural buffer",
    "conservation easement",
    "earthquake zone", "seismic zone", "seismic hazard",
    "tsunami zone", "tsunami hazard", "liquefaction zone",
    "steep slope", "slope hazard", "landslide zone", "geologic hazard",
    # Regional
    "npdes buffer", "flood administration", "view corridor", "viewshed",
    "water rights overlay", "dark sky", "ceqa overlay",
]

# Cluster F: Cadastral and Property Reference
CLUSTER_F = [
    "parcels", "parcel", "assessor parcels", "assessor parcel", "pid",
    "tif parcels", "cfd parcels", "bmr parcels", "flood zone parcels",
    "religious parcel", "building footprints landuse", "lot",
    "parcel owner", "parcel report", "innovation parcel", "right of way",
    "section township range", "address",
]

# Cluster G: Administrative Boundaries
CLUSTER_G = [
    "neighborhoods", "neighborhood", "neighborhood boundary",
    "neighborhood association",
    "city limits", "city boundary", "municipal boundary", "municipalities",
    "county boundary", "adjacent counties",
    "council districts", "wards", "council wards",
    "zip codes", "zipcodes", "zip code boundaries",
    "census tracts", "voting districts", "election districts",
    "school districts", "school catchment", "school attendance zone",
    "school district", "elementary school district",
    "middle school district", "high school district",
    "township", "townships",
    "etj", "extraterritorial jurisdiction",
    "sphere of influence", "urban growth boundary",
    "metro",
    "public owned land",
    "fire districts", "fire response districts", "fire tax districts",
    "greenway",
    "planning area",
]

# Cluster H: Regulatory Use Restrictions
CLUSTER_H = [
    "billboard buffer", "billboards exclusionary zone",
    "outdoor lighting code", "short term rental", "prohibited area",
    "resort hotels", "small cell wireless", "cell towers",
    "scenic byway", "pedestrian mall", "gaming overlay", "gaming district",
    "community residence", "symphony district", "alcohol buffer",
    "adult use overlay", "cannabis overlay", "noise contour",
    "military influence area",
]

# Cluster I: Hazards and Development Restrictions
CLUSTER_I = [
    "floodplain", "flood plain", "flood zone", "floodway", "floodways",
    "floodway channels", "floodway creeks", "flood area",
    "100 year flood", "100-year flood", "100 year floodway",
    "500 year flood", "500-year flood",
    "fema flood", "fema floodplain", "fema firm",
    "special flood hazard area", "sfha", "firm",
    "fire zone", "fire hazard", "fire buffer",
    "wildfire", "wildfire district", "wildfire zone", "wildfire hazard", "wui",
    "earthquake zone", "seismic zone", "seismic hazard",
    "liquefaction", "liquefaction zone", "tsunami zone", "tsunami hazard",
    "geologic hazard",
    "steep slope", "slope restriction", "slope hazard", "slope overlay",
    "landslide", "landslide zone", "erosion zone",
    "contour", "contours", "10 foot contour", "100 foot contour", "topo contour",
    "wetlands", "wetland buffer", "wetland setback",
    "riparian buffer", "riparian zone", "riparian corridor",
    "stream buffer", "stream setback", "river buffer", "waterbody setback",
    "rivers", "streams", "ponds", "lakes", "waterbodies", "water bodies",
    "watersheds", "hydrology",
    "conservation easement", "easement",
    "agricultural district", "voluntary agricultural",
    "voluntary agricultural parcels half mile buffer",
    "right to farm", "rtf zone", "rtf district", "farm buffer",
    "non attainment",
]

# Cluster J: Landmarks and Civic Features (supporting only — requires co-occurring signal)
CLUSTER_J = [
    "high school", "middle school", "elementary school", "public school",
    "school location", "school point", "school site",
    "parks", "park location", "park point",
    "civic area", "civic center", "city hall", "county hall", "county courthouse", "courthouse",
    "fire station", "police station", "library", "community center",
    "transit stop", "transit node", "bus stop", "light rail stop", "train station",
    "commuter rail", "airport",
    "rivers", "streams", "lakes", "lakefront", "waterfront", "water feature",
    "fountain", "playground", "ball field", "athletic field", "stadium", "sports complex",
    "recreation area", "open space", "greenway",
]

# Combined flat list for simple keyword matching (backward compat)
LAYER_KEYWORDS = (
    CLUSTER_A + CLUSTER_B + CLUSTER_C + CLUSTER_D +
    CLUSTER_E + CLUSTER_F + CLUSTER_G + CLUSTER_H +
    CLUSTER_I + CLUSTER_J
)

# ---------------------------------------------------------------------------
# Exclusion Signals — High-Confidence Non-Planning Indicators
# ---------------------------------------------------------------------------

# Service path exclusion tokens
EXCLUDE_SERVICE_TOKENS = [
    "pw_", "scl", "rtc_", "bus_stops", "live_scl", "dpw", "publicworks",
    "fd_", "phantoms", "fire_", "ems_",
    "live_clv_bus", "clv_bus", "business_license",
    "parks_protected", "pools_view", "community_centers_view",
    "utilities", "water_", "sewer_", "storm_", "stormwater_",
    "recycled_water", "lucity",
    "ccsd_schools", "privateschools",
    "police_", "mark43_",
]

# Layer name exclusion keywords
EXCLUDE_LAYER_KEYWORDS = [
    # Public Works / Transportation
    "streets", "major streets", "bike trails", "bicycle lane", "bicycle route",
    "equestrian trails", "trail crossing", "trailheads", "trail projects",
    "trails network", "bus stops", "pavement condition", "pavement index",
    "street sweeping", "guardrails", "handicap ramps", "street lights",
    "traffic signal", "truck routes", "crossroads", "street centerlines",
    "road closures",
    # Fire / Emergency (note: "fire districts", "fire response districts",
    # "fire tax districts" are planning reference boundaries — NOT excluded here.
    # They are excluded only when in a Fire service folder via service-path exclusion.)
    "fire map", "phantoms", "emergency", "ems",
    "fire pre plan", "fire run", "fire stations", "fire incidents",
    "fire hydrants", "fire water points",
    # Business License
    "business licenses", "active business licenses",
    "gaming restricted", "gaming non-restricted",
    "alcohol on-premise", "alcohol off-premise",
    "massage establishment", "marijuana establishments",
    "daily labor service", "financial institution",
    "amusement park", "open air vending",
    "alcohol beverage control",
    # Parks (standalone)
    "pools", "community centers", "park lights", "park pathways",
    "park points",
    # Utilities
    "water hydrant", "water meter", "water service", "sewer", "sanitary",
    "storm drain", "recycled water", "irrigation controller",
    "backflow", "odor sample",
    # Schools (standalone)
    "ccsd schools", "private schools", "school points",
    # Police
    "police beats", "police traffic citations", "crime",
    "tiburon reporting districts",
    # Transportation
    "airports", "roads", "railroads",
]

# Raster/non-vector exclusion patterns
EXCLUDE_PATTERNS = [
    re.compile(r"\b(raster|image|imagery|aerial|lidar|ortho|dem|elevation|hillshade|basemap|tile|cache)\b", re.IGNORECASE),
]


# Path to the planning layer pattern skill document (used by LLM search prompt
# and as the canonical source for keyword lists)
_SKILL_DOC_PATH = os.path.join(os.path.dirname(__file__), "docs", "planning-layer-pattern-skill-v2.md")


# ---------------------------------------------------------------------------
# Load keywords from planning-layer-pattern-skill-v2.md
# ---------------------------------------------------------------------------
# The planning doc is the canonical source of truth for all keyword lists.
# The hardcoded lists above are fallbacks in case the doc is missing.

def _parse_backtick_keywords(line: str) -> list[str]:
    """Extract backtick-delimited keywords from a line like:
    Keywords: `Foo`, `Bar Baz`, `Qux` (note text)
    Returns lowercased keywords with regex wildcards (.*) replaced by spaces
    and parenthetical notes stripped.
    """
    keywords = []
    for match in re.finditer(r"`([^`]+)`", line):
        kw = match.group(1)
        # Strip parenthetical notes like "(when in planning context)"
        kw = re.sub(r"\s*\(.*?\)\s*", "", kw).strip()
        # Replace regex wildcards with space for substring matching
        kw = kw.replace(".*", " ").strip()
        if kw and len(kw) > 1:
            keywords.append(kw.lower())
    return keywords


def _load_keywords_from_doc():
    """Parse the planning doc and update module-level keyword lists.
    Called once at module load time."""
    global CLUSTER_A, CLUSTER_B, CLUSTER_C, CLUSTER_D, CLUSTER_E
    global CLUSTER_F, CLUSTER_G, CLUSTER_H, CLUSTER_I, CLUSTER_J
    global LAYER_KEYWORDS
    global EXCLUDE_SERVICE_TOKENS, EXCLUDE_LAYER_KEYWORDS

    if not os.path.isfile(_SKILL_DOC_PATH):
        return

    try:
        with open(_SKILL_DOC_PATH, "r", encoding="utf-8") as f:
            md = f.read()
    except OSError:
        return

    cluster_map = {
        "cluster a": "CLUSTER_A", "cluster b": "CLUSTER_B",
        "cluster c": "CLUSTER_C", "cluster d": "CLUSTER_D",
        "cluster e": "CLUSTER_E", "cluster f": "CLUSTER_F",
        "cluster g": "CLUSTER_G", "cluster h": "CLUSTER_H",
        "cluster i": "CLUSTER_I", "cluster j": "CLUSTER_J",
    }

    current_cluster = None
    current_excl_section = None
    excl_svc: list[str] = []
    excl_lyr: list[str] = []

    for line in md.splitlines():
        stripped = line.strip()

        # Detect cluster headings like "### Cluster A: Zoning..."
        if stripped.startswith("### Cluster"):
            for key, attr in cluster_map.items():
                if key in stripped.lower():
                    current_cluster = attr
                    current_excl_section = None
                    break
            continue

        # Detect exclusion subsection headings
        if stripped.startswith("### ") and current_cluster is None:
            lower = stripped.lower()
            if any(x in lower for x in ["public works", "fire / emergency",
                    "business license", "parks and recreation", "utilities",
                    "police", "schools"]):
                current_excl_section = stripped
                continue

        # Detect end of section (new ## heading)
        if stripped.startswith("## "):
            current_cluster = None
            current_excl_section = None
            continue

        # Parse Keywords: lines for clusters — merge with hardcoded defaults
        # so the doc adds keywords but manually-added ones aren't lost
        if current_cluster and stripped.lower().startswith("keywords:"):
            kws = _parse_backtick_keywords(stripped)
            if kws:
                existing = set(globals().get(current_cluster, []))
                merged = list(existing | set(kws))
                globals()[current_cluster] = merged
            current_cluster = None
            continue

        # Parse exclusion sections
        if current_excl_section:
            lower = stripped.lower()
            if lower.startswith("service name tokens:"):
                for kw in _parse_backtick_keywords(stripped):
                    excl_svc.append(kw)
            elif lower.startswith("layer name keywords:"):
                for kw in _parse_backtick_keywords(stripped):
                    excl_lyr.append(kw)

    # Update module globals
    if excl_svc:
        EXCLUDE_SERVICE_TOKENS = excl_svc
    if excl_lyr:
        EXCLUDE_LAYER_KEYWORDS = excl_lyr

    LAYER_KEYWORDS = (
        CLUSTER_A + CLUSTER_B + CLUSTER_C + CLUSTER_D +
        CLUSTER_E + CLUSTER_F + CLUSTER_G + CLUSTER_H +
        CLUSTER_I + CLUSTER_J
    )


# Load from doc at module import time
_load_keywords_from_doc()


# ---------------------------------------------------------------------------
# Progress tracker (console-based for intranet/headless use)
# ---------------------------------------------------------------------------

class ProgressTracker:
    """Progress reporter that supports both console and callback modes."""

    def __init__(self, callback=None):
        self.steps: list[str] = []
        self.stats: dict[str, int] = defaultdict(int)
        self._callback = callback

    def log(self, message: str):
        self.steps.append(message)
        print(f"  [*] {message}")
        if self._callback:
            self._callback("log", message)

    def stat(self, key: str, value: int):
        self.stats[key] = value
        if self._callback:
            self._callback("stat", f"{key}: {value}")

    def summary(self):
        print("\n" + "=" * 60)
        print("  SCAN SUMMARY")
        print("=" * 60)
        for k, v in self.stats.items():
            print(f"  {k}: {v}")
        print("=" * 60 + "\n")
        if self._callback:
            self._callback("summary", json.dumps(dict(self.stats)))

    def reset(self, callback=None):
        """Reset state for a new scan run."""
        self.steps.clear()
        self.stats.clear()
        self._callback = callback


# Module-level default instance (used by CLI mode)
progress = ProgressTracker()


# ---------------------------------------------------------------------------
# HTTP helpers
# ---------------------------------------------------------------------------

session = requests.Session()
session.headers.update(HEADERS)


def fetch(url: str, timeout: int = REQUEST_TIMEOUT) -> Optional[requests.Response]:
    """GET with retries and error handling."""
    last_error = None
    for attempt in range(3):
        try:
            r = session.get(url, timeout=timeout, allow_redirects=True)
            r.raise_for_status()
            return r
        except requests.exceptions.SSLError:
            # Retry once without SSL verification for government sites
            # with misconfigured certificates
            try:
                r = session.get(url, timeout=timeout, allow_redirects=True,
                                verify=False)
                r.raise_for_status()
                return r
            except requests.RequestException as e:
                last_error = f"SSL error and fallback failed: {e}"
                break
        except requests.exceptions.HTTPError as e:
            last_error = f"HTTP {r.status_code}"
            # Some government pages return 403 but still have usable content
            if r.status_code in (403, 406):
                return r
            if attempt < 2:
                time.sleep(1 * (attempt + 1))
        except requests.exceptions.ConnectionError as e:
            last_error = f"Connection error: {e}"
            if attempt < 2:
                time.sleep(1 * (attempt + 1))
        except requests.exceptions.Timeout:
            last_error = "Request timed out"
            if attempt < 2:
                time.sleep(1 * (attempt + 1))
        except requests.RequestException as e:
            last_error = str(e)
            if attempt < 2:
                time.sleep(1 * (attempt + 1))
    if last_error:
        progress.log(f"    ⚠ Fetch failed: {last_error}")
    return None


# ---------------------------------------------------------------------------
# URL validation
# ---------------------------------------------------------------------------

def validate_url(url: str, mode: str = "homepage") -> dict:
    """
    Validate a URL before starting a scan.
    Returns dict with: valid (bool), message (str), status_code (int or None).
    """
    # Basic format check
    parsed = urlparse(url)
    if not parsed.scheme:
        return {"valid": False, "message": "URL must start with http:// or https://",
                "status_code": None}
    if parsed.scheme not in ("http", "https"):
        return {"valid": False, "message": f"Unsupported scheme '{parsed.scheme}'. Use http:// or https://",
                "status_code": None}
    if not parsed.netloc:
        return {"valid": False, "message": "URL has no domain name.",
                "status_code": None}

    # For direct mode, check that URL looks like an ArcGIS REST endpoint
    if mode == "direct":
        if "/rest/services" not in url.lower():
            return {"valid": False,
                    "message": "For Direct mode, URL should contain '/rest/services' "
                               "(e.g. https://gis.example.gov/arcgis/rest/services)",
                    "status_code": None}

    # Reachability check
    try:
        r = session.head(url, timeout=10, allow_redirects=True)
        if r.status_code < 400:
            return {"valid": True,
                    "message": f"URL is reachable (HTTP {r.status_code}).",
                    "status_code": r.status_code}
        # Some government sites block HEAD requests; retry with GET
        r = session.get(url, timeout=10, allow_redirects=True)
        if r.status_code < 400:
            return {"valid": True,
                    "message": f"URL is reachable (HTTP {r.status_code}).",
                    "status_code": r.status_code}
        return {"valid": False,
                "message": f"URL returned HTTP {r.status_code}.",
                "status_code": r.status_code}
    except requests.exceptions.SSLError:
        # Retry without SSL verification for misconfigured government certs
        try:
            r = session.head(url, timeout=10, allow_redirects=True, verify=False)
            if r.status_code < 400:
                return {"valid": True,
                        "message": f"URL is reachable (HTTP {r.status_code}, SSL certificate issue bypassed).",
                        "status_code": r.status_code}
            return {"valid": False,
                    "message": f"URL returned HTTP {r.status_code} (SSL certificate issue bypassed).",
                    "status_code": r.status_code}
        except requests.RequestException as e:
            return {"valid": False,
                    "message": f"SSL certificate error and fallback failed: {e}",
                    "status_code": None}
    except requests.ConnectionError:
        return {"valid": False, "message": "Could not connect to the server. Check the URL and your network.",
                "status_code": None}
    except requests.Timeout:
        return {"valid": False, "message": "Connection timed out. The server may be down.",
                "status_code": None}
    except requests.RequestException as e:
        return {"valid": False, "message": f"Request failed: {e}",
                "status_code": None}


# ---------------------------------------------------------------------------
# Interaction helper (for prompting user mid-scan)
# ---------------------------------------------------------------------------

class InteractionRequest:
    """Allows the scanner to pause and ask the user a question."""

    def __init__(self):
        self._response_event = threading.Event()
        self._response_value: Optional[str] = None

    def ask(self, prompt_callback, question: str, options: list[str]) -> str:
        """
        Send a prompt to the user via the progress callback and block
        until a response is received.
        """
        self._response_event.clear()
        self._response_value = None
        if prompt_callback:
            prompt_callback("prompt", json.dumps({
                "question": question,
                "options": options,
            }))
        # Block until user responds (timeout after 5 minutes)
        self._response_event.wait(timeout=300)
        return self._response_value or options[0]  # default to first option

    def respond(self, value: str):
        """Called from the web layer when the user answers."""
        self._response_value = value
        self._response_event.set()


# ---------------------------------------------------------------------------
# Step 1 – Crawl government site to find ArcGIS REST endpoints
# ---------------------------------------------------------------------------

def extract_arcgis_rest_urls(text: str) -> set[str]:
    """Pull ArcGIS REST Services Directory URLs from raw text."""
    urls: set[str] = set()
    for pat in ARCGIS_REST_PATTERNS:
        for m in pat.finditer(text):
            url = m.group(0).rstrip("/")
            # Normalise to the root services directory
            idx = url.lower().find("/rest/services")
            if idx != -1:
                url = url[: idx + len("/rest/services")]
            urls.add(url)
    return urls


def _normalize_rest_directory(url: str) -> str:
    """
    Given a URL that may point to a specific service (e.g.
    .../rest/services/Addressing/FeatureServer), traverse up to the
    root services directory (.../rest/services).
    """
    idx = url.lower().find("/rest/services")
    if idx == -1:
        return url
    return url[: idx + len("/rest/services")]


def expand_single_service_urls(found_urls: set[str]) -> set[str]:
    """
    If a discovered URL points to a single FeatureServer/MapServer (e.g.
    .../rest/services/Addressing/FeatureServer), traverse up to the root
    services directory so we enumerate all services, not just one.
    """
    expanded: set[str] = set()
    for url in found_urls:
        root = _normalize_rest_directory(url)
        expanded.add(root)
        if root != url:
            progress.log(f"  Expanded single service URL to directory root: {root}")
    return expanded


def guess_arcgis_urls(start_url: str) -> set[str]:
    """Tier 1 — Fast-path probe: test common ArcGIS hosting patterns.

    Constructs candidate URLs from common subdomain patterns (gis., maps.,
    etc.) and ArcGIS Online services patterns, then tests each with a
    lightweight HTTP request.  This is free (no API cost) and fast.

    UPDATE: Add or remove candidate patterns below to match hosting
    conventions used by your target jurisdictions.
    """
    parsed = urlparse(start_url)
    domain_parts = parsed.netloc.replace("www.", "").split(".")
    city_slug = domain_parts[0] if domain_parts else ""

    # Strip "www." from netloc for subdomain probing
    base_domain = parsed.netloc.replace("www.", "")

    candidates = [
        f"https://gis.{base_domain}/arcgis/rest/services",
        f"https://maps.{base_domain}/arcgis/rest/services",
        f"https://mapping.{base_domain}/arcgis/rest/services",
        f"https://gisweb.{base_domain}/arcgis/rest/services",
        f"https://services.{base_domain}/arcgis/rest/services",
        f"https://arcgis.{base_domain}/arcgis/rest/services",
        f"https://webgis.{base_domain}/arcgis/rest/services",
        f"https://egis.{base_domain}/arcgis/rest/services",
        f"https://{parsed.netloc}/arcgis/rest/services",
    ]
    # Note: ArcGIS Online (services1-9.arcgis.com) org IDs are random
    # alphanumeric strings, not derivable from the jurisdiction name.
    # AGOL discovery requires the LLM search (Tier 2).

    found: set[str] = set()
    for url in candidates:
        resp = fetch(url, timeout=10)
        if resp is None or resp.status_code != 200:
            continue
        # Validate that the endpoint actually contains services or folders —
        # some ArcGIS servers return 200 with an empty services page
        try:
            data = resp.json()
            has_content = data.get("services") or data.get("folders")
        except (json.JSONDecodeError, ValueError):
            # Not JSON — check for HTML services directory page
            has_content = "rest/services" in resp.text.lower() and (
                "MapServer" in resp.text or "FeatureServer" in resp.text
            )
        if has_content:
            progress.log(f"  Guessed valid endpoint: {url}")
            found.add(url)
    return found


# ---------------------------------------------------------------------------
# Tier 1.5 — ArcGIS Online (AGOL) content search
# ---------------------------------------------------------------------------
# Esri's public content search API lets us find a jurisdiction's GIS data
# by searching for common layer names (zoning, parcels, land use) combined
# with the jurisdiction name. This is free, deterministic, and more reliable
# than LLM web search for AGOL-hosted jurisdictions.

_AGOL_SEARCH_URL = "https://www.arcgis.com/sharing/rest/search"

# Common planning layer names to search for — these are near-universal
_AGOL_SEARCH_TERMS = ["zoning", "parcels", "land use"]


def _search_agol_for_jurisdiction(jurisdiction_name: str) -> set[str]:
    """Search ArcGIS Online for a jurisdiction's REST Services Directory.

    Searches AGOL's public content API for Feature Services matching
    the jurisdiction name + common planning keywords. Ranks results by
    how closely the owner name matches the jurisdiction, then extracts
    the REST directory root from the best match.
    """
    from urllib.parse import quote

    # Build search terms from jurisdiction name
    # Strip "City of", "County of" etc. for cleaner owner matching
    name_words = [w for w in jurisdiction_name.lower().split()
                  if w not in ("city", "of", "county", "town", "village", "the")]
    name_key = "".join(name_words)  # e.g., "milpitas", "lasvegas"

    owner_urls: dict[str, list[str]] = {}

    for term in _AGOL_SEARCH_TERMS:
        query = f'{jurisdiction_name} {term} type:"Feature Service"'
        try:
            resp = session.get(
                _AGOL_SEARCH_URL,
                params={"q": query, "f": "json", "num": 10},
                timeout=10,
            )
            if resp.status_code != 200:
                continue
            data = resp.json()
        except (requests.RequestException, json.JSONDecodeError, ValueError):
            continue

        for item in data.get("results", []):
            owner = item.get("owner", "")
            url = item.get("url", "")
            if not owner or not url:
                continue
            if owner not in owner_urls:
                owner_urls[owner] = []
            owner_urls[owner].append(url)

    if not owner_urls:
        progress.log("  AGOL search returned no results")
        return set()

    # Build abbreviation variants for matching
    # "Grand Prairie" → "gp", "grandprairie"
    # "Las Vegas" → "lv", "lasvegas"
    initials = "".join(w[0] for w in name_words if w)  # "gp", "lv"

    # Rank owners by how well they match the jurisdiction
    best_owner = None
    best_score = -1

    for owner, urls in owner_urls.items():
        score = len(urls)  # frequency of results
        owner_clean = owner.lower().replace("_", "").replace("-", "").replace(".", "")
        # Also clean the email portion if present
        owner_name = owner_clean.split("@")[0] if "@" in owner_clean else owner_clean

        # Strong signal: full jurisdiction name appears in owner name
        if name_key in owner_name:
            score += 100
        # Partial match on individual words (4+ chars to avoid false positives)
        elif any(w in owner_name for w in name_words if len(w) > 3):
            score += 50
        # Abbreviation match: "gpgis" contains "gp" for "Grand Prairie"
        # Require initials to be at least 2 chars and owner to contain "gis"
        elif len(initials) >= 2 and initials in owner_name and "gis" in owner_name:
            score += 60
        # GIS-related owner names are more likely to be official
        if "gis" in owner_name:
            score += 10
        # Email domain matching: gpgis@cityofgp.com → "cityofgp" contains "gp"
        if "@" in owner_clean:
            email_domain = owner_clean.split("@")[1].split(".")[0]
            if name_key in email_domain or (len(initials) >= 2 and
                    ("cityof" + initials) == email_domain):
                score += 80

        if score > best_score:
            best_score = score
            best_owner = owner

    if best_score < 50:
        # No confident match — jurisdiction name not found in any owner
        progress.log(f"  AGOL search found results but no confident jurisdiction match "
                     f"(best owner: {best_owner}, score: {best_score})")
        return set()

    urls = owner_urls[best_owner]
    progress.log(f"  Found AGOL owner: {best_owner} ({len(urls)} layers)")

    # Extract REST directory roots from the URLs
    roots: set[str] = set()
    for url in urls:
        root = _normalize_rest_directory(url)
        if root:
            roots.add(root)

    for root in roots:
        progress.log(f"  ✓ {root}")

    return roots


# ---------------------------------------------------------------------------
# Step 1B – LLM web search for ArcGIS REST endpoints
# ---------------------------------------------------------------------------

# ---------------------------------------------------------------------------
# Claude API configuration (Tier 2 — LLM Web Search)
# ---------------------------------------------------------------------------
# SECURITY: The API key is read from the environment variable at startup.
# It is NEVER logged, printed to console, or included in output files.
# To set the key:
#   Windows PowerShell:  $env:ANTHROPIC_API_KEY = "sk-ant-..."
#   Linux / macOS:       export ANTHROPIC_API_KEY="sk-ant-..."
# Generate your key at: https://console.anthropic.com → API Keys
_ANTHROPIC_API_KEY = os.environ.get("ANTHROPIC_API_KEY", "")

# Claude model used for web search.  Override with env var ARCGIS_SCANNER_MODEL.
# Default is claude-sonnet-4-6 (cost-effective for search tasks).
_LLM_MODEL = os.environ.get("ARCGIS_SCANNER_MODEL", "claude-sonnet-4-6")

def _build_llm_search_system_prompt() -> str:
    """Build the LLM search system prompt, enriched with context from the
    planning-layer-pattern-skill-v2.md document.  This gives the LLM
    domain-specific vocabulary and URL structure patterns so it can craft
    better search queries using the jurisdiction's name and URL reference."""

    # Load planning doc context for search guidance
    planning_context = ""
    if os.path.isfile(_SKILL_DOC_PATH):
        try:
            with open(_SKILL_DOC_PATH, "r", encoding="utf-8") as f:
                md = f.read()
            # Extract the key sections the LLM needs for search context:
            # URL structure patterns, Tier 1 tokens, and abbreviations
            sections: list[str] = []
            current_section = ""
            capture = False
            target_headings = [
                "## 2. Tier 1",           # Service path signals
                "## 6. URL Structure",     # URL patterns
                "## 7. Named Abbreviations", # Acronyms
                "## 10. Regional Terminology", # Regional terms
            ]
            for line in md.splitlines():
                if line.strip().startswith("## "):
                    if capture and current_section:
                        sections.append(current_section)
                    capture = any(h in line for h in target_headings)
                    current_section = line + "\n" if capture else ""
                elif capture:
                    current_section += line + "\n"
            if capture and current_section:
                sections.append(current_section)
            if sections:
                planning_context = "\n".join(sections)
        except OSError:
            pass  # fall back to base prompt without planning context

    base_prompt = """\
You are a web search agent. Your sole task is to find the ArcGIS REST Services \
Directory URL for a specific local government jurisdiction.

## What You Are Looking For

You are searching for EITHER of these two things:

### Option 1: ArcGIS REST Services Directory URL (preferred)
A URL containing one of these patterns:
- /arcgis/rest/services
- /ArcGIS/rest/services
- /server/rest/services
- rest/services

Typically hosted at subdomains like:
- https://gis.<domain>/arcgis/rest/services
- https://maps.<domain>/arcgis/rest/services
- https://mapping.<domain>/arcgis/rest/services
- https://gisweb.<domain>/arcgis/rest/services
- https://services.<domain>/arcgis/rest/services
- https://arcgis.<domain>/arcgis/rest/services
- https://webgis.<domain>/arcgis/rest/services
- https://egis.<domain>/arcgis/rest/services
- https://geoportal.<domain>/arcgis/rest/services
- https://opendata.<domain>/arcgis/rest/services
- https://<jurisdiction>.maps.arcgis.com
- https://<jurisdiction>.hub.arcgis.com

### Option 2: Individual Feature Layer URL (also acceptable!)
If you cannot find the REST Services Directory root, finding ANY individual \
FeatureServer or MapServer URL is also valuable. These look like:
- https://services8.arcgis.com/<orgid>/ArcGIS/rest/services/<LayerName>/FeatureServer
- https://gis.<domain>/arcgis/rest/services/<Folder>/<ServiceName>/MapServer
- https://maps.<domain>/arcgis/rest/services/<ServiceName>/FeatureServer

For example, searching for a jurisdiction's "Zoning" or "Parcels" layer might \
return a URL like:
  https://services8.arcgis.com/OPmRdssd8jj0bT5H/ArcGIS/rest/services/Zoning/FeatureServer

This is useful because we can derive the REST directory root by truncating \
at "/services/".

### Secondary signals (promising but not sufficient alone)
- arcgis.com/home, /arcgis/home, /portal/home
- hub.arcgis.com, opendata.arcgis.com
- /apps/webappviewer, /apps/mapviewer
- "Powered by Esri", "Built with ArcGIS"
- MapServer, FeatureServer, ImageServer, GeocodeServer, GPServer

## Instructions

1. Search the web using the jurisdiction's FULL NAME and its WEBSITE DOMAIN \
as primary search parameters.  For example, if the jurisdiction is "City of Dublin" \
and the domain is "dublinohiousa.gov", search for:
   - "dublinohiousa.gov arcgis rest services"
   - "City of Dublin Ohio GIS map services"
   - "dublinohiousa.gov GIS"
   - "site:dublinohiousa.gov arcgis"
2. Also try domain-variant searches using common GIS subdomains:
   - "gis.dublinohiousa.gov", "maps.dublinohiousa.gov"
3. Try multiple search queries if the first doesn't find it (e.g., include "GIS", \
"ArcGIS", "rest services", "map services", the jurisdiction name).
4. Search for the jurisdiction's planning/community development GIS pages, as \
these often link to the ArcGIS REST directory.  Use department names like \
"Planning", "Community Development", "ComDev", "CDD", "Land Development" in queries.
5. If you STILL haven't found the REST directory, search for specific feature \
layer names combined with the jurisdiction name. Try:
   - "<jurisdiction> zoning FeatureServer"
   - "<jurisdiction> parcels arcgis"
   - "<jurisdiction> land use MapServer"
   - "<jurisdiction> planning GIS layers"
These searches often surface individual feature layer URLs hosted on \
services.arcgis.com or the jurisdiction's GIS server.
6. Examine search results for URLs matching the target patterns above.
7. If you find a matching URL, verify it by fetching the page if possible.
8. Report ALL matching URLs you find — both REST directory roots AND individual \
FeatureServer/MapServer layer URLs.

## Output Format

After your search, respond with ONLY a JSON object (no markdown fences):
{
  "found": true or false,
  "urls": ["https://gis.example.gov/arcgis/rest/services", ...],
  "confidence": "confirmed" or "probable" or "not_found",
  "notes": "Brief explanation of what you found"
}

Include ALL URLs you found — both REST directory roots and individual feature \
layer URLs (FeatureServer, MapServer). We will extract the directory root from \
feature layer URLs automatically.

If you find NO matching URLs, set "found" to false and "urls" to an empty list.
Do NOT fabricate or guess URLs. Only report URLs you actually found in search results."""

    if planning_context:
        base_prompt += f"""

## Planning Layer Pattern Reference (from planning-layer-pattern-skill-v2.md)

Use the following reference to understand common URL structures, service path \
tokens, abbreviations, and regional terminology for government GIS services. \
This will help you craft better search queries for the jurisdiction.

{planning_context}"""

    return base_prompt


_LLM_SEARCH_SYSTEM_PROMPT = _build_llm_search_system_prompt()


def llm_search_for_arcgis(jurisdiction_name: str, homepage_url: str = "") -> set[str]:
    """
    Two-Tier ArcGIS endpoint discovery.

    This is the main entry point for finding ArcGIS REST Services Directory
    URLs when the user provides a jurisdiction name/homepage (Path B).

    The LLM searches for EITHER the REST Services Directory URL OR individual
    feature layer URLs (FeatureServer/MapServer).  When a feature layer URL
    is found, the REST directory root is derived by truncating at '/services/'.

    Execution order:
      1. FAST-PATH PROBE (Tier 1) — tests common subdomain patterns like
         gis.<domain>, maps.<domain> with HTTP requests.  Free, no API cost.
         If found, returns immediately.
      2. LLM WEB SEARCH (Tier 2) — calls Claude API with the web_search
         server tool.  Requires ANTHROPIC_API_KEY env var and the
         ``anthropic`` Python package.

    Args:
        jurisdiction_name: Human-readable name (e.g. "City of Las Vegas").
            Sanitised internally before use.
        homepage_url: The jurisdiction's website URL.

    Returns:
        Set of discovered ArcGIS REST Services Directory root URLs.
    """
    # SECURITY: Sanitise jurisdiction name to prevent prompt injection.
    # Only alphanumeric, spaces, hyphens, periods, apostrophes are allowed.
    jurisdiction_name = sanitize_jurisdiction_name(jurisdiction_name)

    # --- Tier 1: Fast-path probe (FREE, no API cost) ---
    # Tests common subdomain patterns (gis., maps., etc.) directly.
    # If any return a valid ArcGIS services page, we skip everything else.
    if homepage_url:
        progress.log("Tier 1 — Fast-path: probing common ArcGIS subdomains…")
        fast_results = guess_arcgis_urls(homepage_url)
        if fast_results:
            progress.log(f"Tier 1 — Fast-path found {len(fast_results)} endpoint(s)")
            return expand_single_service_urls(fast_results)

    # --- Tier 1.5: ArcGIS Online search (FREE, no API cost) ---
    # Search Esri's public AGOL content API for the jurisdiction's org.
    # This is the most reliable method for AGOL-hosted jurisdictions.
    progress.log("Tier 1.5 — Searching ArcGIS Online for jurisdiction data…")
    agol_results = _search_agol_for_jurisdiction(jurisdiction_name)
    if agol_results:
        progress.log(f"Tier 1.5 — AGOL search found {len(agol_results)} endpoint(s)")
        return expand_single_service_urls(agol_results)

    # --- Tier 2: LLM Web Search (requires API key) ---
    if not _ANTHROPIC_API_KEY:
        progress.log("Tier 2 — No ANTHROPIC_API_KEY set. Cannot perform LLM web search.")
        progress.log("Set the ANTHROPIC_API_KEY environment variable, or use "
                      "Direct mode (Path A) with a known ArcGIS REST Services Directory URL.")
        return set()

    try:
        import anthropic
    except ImportError:
        progress.log("ERROR: anthropic package not installed. Run: pip install anthropic")
        return set()

    # --- Tier 2: LLM Web Search (requires API key) ---
    # Rebuild system prompt fresh so it picks up any planning doc changes
    search_system_prompt = _build_llm_search_system_prompt()
    progress.log(f"Tier 2 — LLM web search for ArcGIS REST services: {jurisdiction_name}")

    # Build the search prompt — jurisdiction name and URL are the primary
    # search parameters, per planning-layer-pattern-skill-v2.md guidance.
    domain_hint = ""
    domain_name = ""
    if homepage_url:
        parsed = urlparse(homepage_url)
        domain_name = parsed.netloc
        domain_hint = (
            f"\nTheir website domain is: {domain_name}\n"
            f"Use this domain as a primary search parameter. Try searches like:\n"
            f'  - "{domain_name} arcgis rest services"\n'
            f'  - "site:{domain_name} arcgis"\n'
            f'  - "gis.{domain_name}" and "maps.{domain_name}"\n'
            f'  - "{jurisdiction_name} GIS map services"\n'
            f'  - "{jurisdiction_name} planning community development GIS"'
        )

    user_prompt = (
        f"Find the ArcGIS REST Services Directory for: {jurisdiction_name}\n"
        f"{domain_hint}\n\n"
        f"IMPORTANT: The results MUST be for {jurisdiction_name} specifically — "
        f"not a neighboring city, county, or regional agency. Verify that any "
        f"ArcGIS endpoint you find is operated by or contains data for "
        f"{jurisdiction_name} before reporting it.\n\n"
        f"Use the jurisdiction name '{jurisdiction_name}' and "
        f"{'domain ' + repr(domain_name) if domain_name else 'common government domain patterns'} "
        f"as your primary search parameters.\n\n"
        f"Search strategy:\n"
        f"1. Search for the ArcGIS REST Services Directory URL (contains "
        f"'/arcgis/rest/services' or similar pattern).\n"
        f"2. Many jurisdictions host GIS data on ArcGIS Online (AGOL) at URLs like "
        f"https://services[N].arcgis.com/[orgId]/ArcGIS/rest/services — search for "
        f"'{jurisdiction_name} arcgis online' or '{jurisdiction_name} GIS hub' "
        f"to find their AGOL organization.\n"
        f"3. Search for their GIS hub page (e.g., {jurisdiction_name.replace(' ', '-').lower()}"
        f"-gis.hub.arcgis.com or similar).\n"
        f"4. If you cannot find the directory root, search for individual "
        f"feature layers:\n"
        f'  - "{jurisdiction_name} zoning FeatureServer"\n'
        f'  - "{jurisdiction_name} parcels arcgis"\n'
        f'  - "{jurisdiction_name} land use MapServer"\n'
        f"Any FeatureServer or MapServer URL will help us locate the directory.\n\n"
        f"Also look for links on their Planning or Community Development "
        f"department pages."
    )

    # SECURITY: The API key is passed directly to the SDK client and is
    # NEVER logged or printed.
    client = anthropic.Anthropic(api_key=_ANTHROPIC_API_KEY)
    messages = [{"role": "user", "content": user_prompt}]

    found_urls: set[str] = set()
    max_continuations = 5

    try:
        for continuation in range(max_continuations + 1):
            progress.log(f"  LLM search call {continuation + 1}…")

            response = client.messages.create(
                model=_LLM_MODEL,
                max_tokens=4096,
                system=search_system_prompt,
                tools=[
                    {"type": "web_search_20250305", "name": "web_search"},
                ],
                messages=messages,
            )

            # If Claude is done, extract the final answer
            if response.stop_reason == "end_turn":
                _parse_llm_search_response(response, found_urls)
                break

            # If Claude needs to continue (server tool hit iteration limit)
            if response.stop_reason == "pause_turn":
                messages = [
                    {"role": "user", "content": user_prompt},
                    {"role": "assistant", "content": response.content},
                ]
                continue

            # Unexpected stop reason
            progress.log(f"  Unexpected stop_reason: {response.stop_reason}")
            _parse_llm_search_response(response, found_urls)
            break

    except Exception as e:
        # SECURITY: Log only the error type/message, never the API key.
        progress.log(f"  Tier 2 — LLM search error: {e}")

    # SECURITY: Validate all URLs returned by the LLM before using them.
    # Remove any that point to private/reserved IP ranges (SSRF protection)
    # or use non-http(s) schemes.
    safe_urls: set[str] = set()
    for url in found_urls:
        if is_safe_url(url):
            safe_urls.add(url)
        else:
            progress.log(f"  Rejected unsafe URL from LLM response: {url[:80]}")
    found_urls = safe_urls

    if found_urls:
        progress.log(f"Tier 2 — LLM search found {len(found_urls)} ArcGIS REST endpoint(s)")
        for url in found_urls:
            progress.log(f"  ✓ {url}")

        # Validate that the discovered endpoints belong to the target jurisdiction
        # by checking if any service names reference the jurisdiction
        validated_urls = _validate_jurisdiction_match(
            found_urls, jurisdiction_name, homepage_url
        )
        if validated_urls:
            found_urls = validated_urls
        else:
            progress.log("  ⚠ Could not confirm jurisdiction match — using all discovered URLs")
    else:
        progress.log("Tier 2 — LLM search did not find any ArcGIS REST endpoints.")
        progress.log("Try Direct mode (Path A) if you can locate the ArcGIS REST Services Directory URL manually.")

    return expand_single_service_urls(found_urls)


def _validate_jurisdiction_match(
    urls: set[str], jurisdiction_name: str, homepage_url: str
) -> set[str] | None:
    """Check if discovered ArcGIS endpoints belong to the target jurisdiction.

    Fetches the services list from each URL and looks for the jurisdiction name
    (or domain slug) in service names, descriptions, or the endpoint URL itself.
    Returns the subset of URLs that match, or None if validation is inconclusive.
    """
    if not jurisdiction_name:
        return None

    # Build search terms from jurisdiction name and domain
    search_terms = []
    # Split jurisdiction name into significant words (skip "City", "of", "County")
    for word in jurisdiction_name.lower().split():
        if word not in ("city", "of", "county", "town", "village", "the"):
            search_terms.append(word)

    if homepage_url:
        parsed = urlparse(homepage_url)
        domain = parsed.netloc.replace("www.", "")
        # Add the main domain slug (e.g., "milpitas" from "milpitas.gov")
        slug = domain.split(".")[0]
        if slug and len(slug) > 2:
            search_terms.append(slug.lower())

    if not search_terms:
        return None

    matched: set[str] = set()
    for url in urls:
        root = _normalize_rest_directory(url)

        # Check if jurisdiction name appears in the URL itself
        url_lower = url.lower()
        if any(term in url_lower for term in search_terms):
            matched.add(url)
            continue

        # Fetch services list and check names/descriptions
        resp = fetch(f"{root}?f=json", timeout=10)
        if resp is None:
            continue
        try:
            data = resp.json()
        except (json.JSONDecodeError, ValueError):
            continue

        # Check service names and folder names
        text_to_check = " ".join(
            svc.get("name", "") for svc in data.get("services", [])
        ).lower()
        text_to_check += " " + " ".join(data.get("folders", [])).lower()
        # Also check the serviceDescription if available
        text_to_check += " " + data.get("serviceDescription", "").lower()

        if any(term in text_to_check for term in search_terms):
            matched.add(url)
            progress.log(f"  ✓ Confirmed jurisdiction match: {root}")
        else:
            progress.log(f"  ✗ No jurisdiction match in services at: {root}")

    return matched if matched else None


def _parse_llm_search_response(response, found_urls: set[str]):
    """Extract ArcGIS REST URLs from the LLM's JSON response.

    SECURITY: URLs extracted here are later validated by ``is_safe_url()``
    in the calling function before any HTTP requests are made to them.
    This two-stage approach ensures untrusted LLM output cannot trigger
    requests to internal/private network addresses.
    """
    for block in response.content:
        if getattr(block, "type", None) != "text":
            continue
        text = block.text.strip()

        # Try to parse as JSON (expected format from the system prompt)
        try:
            data = json.loads(text)
            urls = data.get("urls", [])
            for url in urls:
                if isinstance(url, str) and url.startswith("http"):
                    found_urls.add(url.rstrip("/"))
            if data.get("notes"):
                progress.log(f"  LLM notes: {data['notes']}")
            return
        except (json.JSONDecodeError, ValueError):
            pass

        # Fallback: scan the raw text for ArcGIS REST URL patterns
        rest_urls = extract_arcgis_rest_urls(text)
        if rest_urls:
            found_urls.update(rest_urls)

        # Also scan for individual FeatureServer/MapServer URLs
        feature_pat = re.compile(
            r"https?://[^\s\"'<>]+/rest/services/[^\s\"'<>]+/(?:Feature|Map)Server",
            re.IGNORECASE,
        )
        for m in feature_pat.finditer(text):
            found_urls.add(m.group(0).rstrip("/"))


# ---------------------------------------------------------------------------
# Step 2 – Query ArcGIS REST API for folders, services, and layers
# ---------------------------------------------------------------------------

def query_rest_services(rest_url: str) -> list[dict]:
    """
    Walk the ArcGIS REST Services Directory. Returns a list of dicts:
      {service_name, service_url, layer_name, layer_id, layer_url,
       record_count, geometry_type, service_type}
    """
    progress.log(f"Querying ArcGIS REST directory: {rest_url}")
    layers: list[dict] = []

    # Fetch root
    root = fetch(f"{rest_url}?f=json")
    if root is None:
        progress.log(f"  Could not reach {rest_url}")
        return layers

    try:
        root_data = root.json()
    except (json.JSONDecodeError, ValueError):
        progress.log(f"  Invalid JSON from {rest_url}")
        return layers

    # Collect folders + root-level services
    folders = root_data.get("folders", [])
    services = root_data.get("services", [])

    # Also explore each folder
    for folder in folders:
        folder_url = f"{rest_url}/{folder}?f=json"
        resp = fetch(folder_url)
        if resp:
            try:
                folder_data = resp.json()
                services.extend(folder_data.get("services", []))
            except (json.JSONDecodeError, ValueError):
                pass

    progress.log(f"  Found {len(services)} services across {len(folders)} folders")

    for svc in services:
        svc_name = svc.get("name", "")
        svc_type = svc.get("type", "")

        # We only want FeatureServer or MapServer
        if svc_type not in ("FeatureServer", "MapServer"):
            continue

        svc_url = f"{rest_url}/{svc_name}/{svc_type}"
        resp = fetch(f"{svc_url}?f=json")
        if resp is None:
            continue

        try:
            svc_data = resp.json()
        except (json.JSONDecodeError, ValueError):
            continue

        svc_layers = svc_data.get("layers", [])
        for lyr in svc_layers:
            layer_name = lyr.get("name", "")
            layer_id = lyr.get("id", 0)
            layer_url = f"{svc_url}/{layer_id}"

            # Only store name/URL here — detail fetches are deferred
            # to after filtering to avoid thousands of unnecessary requests
            layers.append({
                "service_name": svc_name,
                "service_url": svc_url,
                "service_type": svc_type,
                "layer_name": layer_name,
                "layer_id": layer_id,
                "layer_url": layer_url,
                "record_count": None,
                "geometry_type": "",
            })

    progress.stat("Total feature layers enumerated", len(layers))
    return layers


# ---------------------------------------------------------------------------
# Step 3 – Filter layers
# ---------------------------------------------------------------------------

def is_excluded_by_pattern(layer_name: str) -> bool:
    """Check raster/imagery exclusion patterns."""
    for pat in EXCLUDE_PATTERNS:
        if pat.search(layer_name):
            return True
    return False


# Section 4.10 — Hard Exclusion Patterns (override all rules)
HARD_EXCLUSION_PATTERNS = [
    # Annotation and label classes — use word boundaries to avoid
    # matching "annexation", "annual", etc.
    re.compile(r"\bannotation", re.IGNORECASE),
    re.compile(r"_anno\b", re.IGNORECASE),
    re.compile(r"^anno_", re.IGNORECASE),
    re.compile(r"\blabels\b", re.IGNORECASE),
    # Platform-internal geometry objects (EnerGov and similar)
    # Covers: History, HistoryPoint, HistoryPolygon, HistoryPolyLine, etc.
    re.compile(r"^History$", re.IGNORECASE),
    re.compile(r"^History\s*(?:Point|Poly|Line|Polygon|PolyLine)", re.IGNORECASE),
    # Covers: Spatial Polyline, Spatial Collection, SpatialCollection*
    re.compile(r"^Spatial\s*(?:Polyline|Collection)", re.IGNORECASE),
    re.compile(r"^SpatialCollection", re.IGNORECASE),
    re.compile(r"^Location$", re.IGNORECASE),
    re.compile(r"^Converted_Graphics", re.IGNORECASE),
    re.compile(r"^Feature\.MAPREAD\.", re.IGNORECASE),
    re.compile(r"^CityWide\.SDE\.", re.IGNORECASE),
    # Generic/ambiguous — A_ prefix is an annotation class naming convention
    re.compile(r"^Default$", re.IGNORECASE),
    re.compile(r"^A_\d+", re.IGNORECASE),
    # Imagery and graphics — word boundaries to avoid false positives
    re.compile(r"\bimagery\b", re.IGNORECASE),
    re.compile(r"\baerial\b", re.IGNORECASE),
    re.compile(r"^Converted_Graphics", re.IGNORECASE),
    # Transit operational layers
    re.compile(r"^Bus_Routes_and_Stops", re.IGNORECASE),
    re.compile(r"^BusStops_", re.IGNORECASE),
    # Street sweeping and signs maintenance
    re.compile(r"Street_Sweeping", re.IGNORECASE),
    re.compile(r"Maintenance Subzones For Signs", re.IGNORECASE),
    # Administrative FEMA / census reference (not substantive planning)
    re.compile(r"Map Index", re.IGNORECASE),
    re.compile(r"Panel Index", re.IGNORECASE),
    # Lead/water quality data
    re.compile(r"^LeadWater_", re.IGNORECASE),
    re.compile(r"^Lead_Copper", re.IGNORECASE),
]


def is_hard_excluded(layer_name: str) -> bool:
    """Section 4.10 — Hard exclusion patterns override ALL rules."""
    for pat in HARD_EXCLUSION_PATTERNS:
        if pat.search(layer_name):
            return True
    return False


def is_service_excluded(service_name: str) -> bool:
    """Check if the service path matches a non-planning service."""
    sn = service_name.lower()
    for token in EXCLUDE_SERVICE_TOKENS:
        if token in sn:
            return True
    # InternalUse services should never be traversed for public planning layers
    if "internaluse" in sn.replace("_", "").replace(" ", ""):
        return True
    return False


def score_service_path(service_name: str) -> int:
    """
    Tier 1 scoring: score based on service/folder path tokens.
    +5 for department identifiers, +4 for land use tokens, +3 for development tokens.
    -5 for exclusion service tokens.
    """
    sn = service_name.lower().replace("/", " ").replace("_", " ").replace("-", " ")
    sn_raw = service_name.lower()
    score = 0

    # Check exclusion service tokens first
    for token in EXCLUDE_SERVICE_TOKENS:
        if token in sn_raw:
            score -= 5
            return score  # strong negative, bail early

    # Department identifiers (+5)
    for token in TIER1_DEPARTMENT_TOKENS:
        if token in sn_raw:
            score = max(score, 5)
            break

    # Land use / zoning tokens (+4)
    if score < 4:
        for token in TIER1_LANDUSE_TOKENS:
            if token in sn_raw:
                score = max(score, 4)
                break

    # Development tokens (+3)
    if score < 3:
        for token in TIER1_DEVELOPMENT_TOKENS:
            if token in sn_raw:
                score = max(score, 3)
                break

    return score


def _count_cluster_hits(text: str, cluster: list[str]) -> int:
    """Count how many keywords from a cluster match in the text."""
    text_lower = text.lower()
    return sum(1 for kw in cluster if kw in text_lower)


def score_layer_name(layer_name: str) -> int:
    """
    Tier 2 scoring: score based on layer name keyword clusters.
    +3 for 3+ keywords from a single cluster, +2 for 1-2 keywords from A-F,
    +1 for administrative geography (Cluster G).
    Negative for exclusion keywords.
    """
    ln = layer_name.lower()

    # Check exclusion keywords
    for kw in EXCLUDE_LAYER_KEYWORDS:
        if kw in ln:
            # Check if it's a standalone match (not nested in a planning term)
            # e.g. "parks" standalone vs "parks master plan"
            has_planning_context = any(pk in ln for pk in [
                "plan", "zone", "zoning", "land use", "overlay", "district",
                "parcel", "subdivision", "historic",
            ])
            if not has_planning_context:
                return -4  # strong negative

    score = 0

    # Score each cluster
    clusters_af = [
        (CLUSTER_A, "A"), (CLUSTER_B, "B"), (CLUSTER_C, "C"),
        (CLUSTER_D, "D"), (CLUSTER_E, "E"), (CLUSTER_F, "F"),
    ]
    max_hits = 0
    total_hits_af = 0
    for cluster, name in clusters_af:
        hits = _count_cluster_hits(ln, cluster)
        max_hits = max(max_hits, hits)
        total_hits_af += hits

    if max_hits >= 3:
        score = max(score, 3)
    elif total_hits_af >= 1:
        score = max(score, 2)

    # Cluster G (administrative geography) — lower confidence
    g_hits = _count_cluster_hits(ln, CLUSTER_G)
    if g_hits > 0 and score == 0:
        score = 1

    # Cluster H (regulatory use restrictions) — moderate confidence
    h_hits = _count_cluster_hits(ln, CLUSTER_H)
    if h_hits > 0:
        score = max(score, 2)

    # Cluster I (hazards and development restrictions) — standalone +2
    i_hits = _count_cluster_hits(ln, CLUSTER_I)
    if i_hits > 0:
        score = max(score, 2)

    # Cluster J (landmarks and civic features) — supporting only, +1
    j_hits = _count_cluster_hits(ln, CLUSTER_J)
    if j_hits > 0 and score == 0:
        score = 1

    return score


def compute_confidence_score(layer: dict) -> int:
    """
    Full confidence score for a layer combining Tier 1 and Tier 2 signals.
    Score >= 4: high confidence, include
    Score 2-3: moderate confidence, include
    Score 0-1: ambiguous
    Score <= -1: exclude
    """
    svc_score = score_service_path(layer["service_name"])
    lyr_score = score_layer_name(layer["layer_name"])
    return svc_score + lyr_score


def filter_layers(layers: list[dict]) -> list[dict]:
    """
    Filter layers using the 3-tier confidence scoring model.
    Includes layers with score >= 2, excludes rasters and low-confidence layers.
    """
    progress.log("Filtering layers using confidence scoring model…")

    scored_layers = []
    excluded_count = 0
    ambiguous_count = 0

    for lyr in layers:
        # Hard exclusion patterns (Section 4.10) — override all rules
        if is_hard_excluded(lyr["layer_name"]):
            excluded_count += 1
            continue

        # Service-path exclusion (InternalUse, Fire_, PW_, etc.)
        if is_service_excluded(lyr["service_name"]):
            excluded_count += 1
            continue

        # Raster/imagery exclusion
        if is_excluded_by_pattern(lyr["layer_name"]):
            excluded_count += 1
            continue

        score = compute_confidence_score(lyr)
        lyr["confidence_score"] = score

        if score >= 4:
            lyr["priority"] = 1  # high confidence
            scored_layers.append(lyr)
        elif score >= 2:
            lyr["priority"] = 2  # moderate confidence
            scored_layers.append(lyr)
        elif score >= 0:
            ambiguous_count += 1
            # Ambiguous — only include if in a planning-named service
            if score_service_path(lyr["service_name"]) >= 3:
                lyr["priority"] = 3
                scored_layers.append(lyr)
        else:
            excluded_count += 1

    # Single-service MapServer rule (planning doc Principle 3):
    # When a generic-named service contains planning layers (score >= 2),
    # promote Cluster G/J layers (score 1) in that same service to score 2.
    # This handles small counties that publish everything in one MapServer.
    services_with_planning = set()
    for lyr in scored_layers:
        if lyr.get("confidence_score", 0) >= 2:
            services_with_planning.add(lyr["service_name"])

    already_included = set(id(lyr) for lyr in scored_layers)
    if services_with_planning and ambiguous_count > 0:
        promoted = 0
        for lyr in layers:
            if id(lyr) in already_included:
                continue
            # Only consider layers that were scored in the first pass
            # (i.e., passed exclusion checks) but scored too low
            if "confidence_score" not in lyr:
                continue
            if lyr["confidence_score"] != 1:
                continue  # only promote score-1 (Cluster G/J) layers
            if lyr["service_name"] not in services_with_planning:
                continue
            lyr["confidence_score"] = 2
            lyr["priority"] = 2
            scored_layers.append(lyr)
            promoted += 1
        if promoted:
            progress.log(f"  Promoted {promoted} administrative layers via single-service rule")

    # Sort by priority then score descending
    scored_layers.sort(key=lambda x: (x.get("priority", 99),
                                       -x.get("confidence_score", 0)))

    high = sum(1 for l in scored_layers if l.get("priority") == 1)
    moderate = sum(1 for l in scored_layers if l.get("priority") == 2)
    low = sum(1 for l in scored_layers if l.get("priority") == 3)

    progress.log(f"  {high} high-confidence, {moderate} moderate-confidence, "
                 f"{low} context-included layers")
    progress.log(f"  {excluded_count} excluded, {ambiguous_count} ambiguous")
    progress.stat("Layers after keyword filter", len(scored_layers))
    return scored_layers


# ---------------------------------------------------------------------------
# Step 4 – Deduplicate
# ---------------------------------------------------------------------------

def normalise_name(name: str) -> str:
    """Lowercase, strip non-alpha, collapse whitespace."""
    n = re.sub(r"[^a-z0-9 ]", " ", name.lower())
    return re.sub(r"\s+", " ", n).strip()


def names_are_similar(a: str, b: str, threshold: float = 0.95) -> bool:
    na, nb = normalise_name(a), normalise_name(b)
    if na == nb:
        return True
    # Strip trailing year-like suffixes so "Parcels2021" matches "Parcels2022"
    na_base = re.sub(r"\d{4}$", "", na).strip()
    nb_base = re.sub(r"\d{4}$", "", nb).strip()
    if na_base and nb_base and na_base == nb_base:
        return True
    # Handle singular/plural: if difference is only a trailing "s"
    if na.rstrip("s") == nb.rstrip("s") and abs(len(na) - len(nb)) <= 1:
        return True
    # Require very high sequence similarity — this prevents collapsing
    # distinct layers like "ZONING" vs "COUNTY ZONING", "PARCELS" vs
    # "VOLUNTARY AGRICULTURAL PARCELS", or "MOUNT AIRY HISTORIC DISTRICT"
    # vs "MOUNT AIRY LOCAL HISTORIC DISTRICT".
    return SequenceMatcher(None, na, nb).ratio() >= threshold


def deduplicate(layers: list[dict]) -> list[dict]:
    """
    Group layers with similar names and keep the best one per group.
    Priority: planning/dev service > highest record count.
    """
    progress.log("Deduplicating layers…")
    groups: list[list[dict]] = []

    for lyr in layers:
        placed = False
        for grp in groups:
            if names_are_similar(lyr["layer_name"], grp[0]["layer_name"]):
                grp.append(lyr)
                placed = True
                break
        if not placed:
            groups.append([lyr])

    deduped: list[dict] = []
    for grp in groups:
        # Sort: priority 1 first, then by record count descending
        grp.sort(key=lambda x: (x.get("priority", 99), -(x.get("record_count") or 0)))
        deduped.append(grp[0])

    removed = len(layers) - len(deduped)
    progress.log(f"  Removed {removed} duplicate layers")
    progress.stat("Layers after deduplication", len(deduped))
    progress.stat("Duplicates removed", removed)
    return deduped


# ---------------------------------------------------------------------------
# Step 4b – Enrich filtered layers with detail (deferred from enumeration)
# ---------------------------------------------------------------------------

def enrich_layer_details(layers: list[dict]):
    """Fetch record count and geometry type for each layer.
    Called after filtering so we only make detail requests for layers
    that passed scoring, not all 1000+ enumerated layers."""
    progress.log(f"Fetching details for {len(layers)} filtered layers…")
    for lyr in layers:
        layer_url = lyr["layer_url"]
        lyr_detail = fetch(f"{layer_url}?f=json")
        if lyr_detail:
            try:
                ld = lyr_detail.json()
                lyr["geometry_type"] = ld.get("geometryType", "")
                count_resp = fetch(f"{layer_url}/query?where=1%3D1&returnCountOnly=true&f=json")
                if count_resp:
                    count_data = count_resp.json()
                    lyr["record_count"] = count_data.get("count")
            except (json.JSONDecodeError, ValueError):
                pass


# ---------------------------------------------------------------------------
# Step 5 – Output
# ---------------------------------------------------------------------------

def write_markdown(layers: list[dict], output_path: str):
    """Write the final table as a Markdown file."""
    lines = [
        "# Government ArcGIS Feature Layers – Scan Results\n",
        "| GIS Layer Name | Source System | Collection | API | Time Period | Update Frequency | Key Data Elements / Notes |",
        "|---|---|---|---|---|---|---|",
    ]
    for lyr in layers:
        name = lyr["layer_name"]
        api_url = lyr["layer_url"]
        lines.append(
            f"| {name} | Esri ArcGIS | API | {api_url} | Current | Ad Hoc | |"
        )
    lines.append("")

    with open(output_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    progress.log(f"Markdown written to {output_path}")


def write_excel(layers: list[dict], output_path: str):
    """Write the final table as an Excel workbook matching the schema."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Feature Layers"

    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")

    headers = [
        "GIS Layer Name",
        "Source System",
        "Collection",
        "API",
        "Time Period",
        "Update Frequency",
        "Key Data Elements / Notes",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True)

    for row_idx, lyr in enumerate(layers, 2):
        ws.cell(row=row_idx, column=1, value=lyr["layer_name"])
        ws.cell(row=row_idx, column=2, value="Esri ArcGIS")
        ws.cell(row=row_idx, column=3, value="API")
        ws.cell(row=row_idx, column=4, value=lyr["layer_url"])
        ws.cell(row=row_idx, column=5, value="Current")
        ws.cell(row=row_idx, column=6, value="Ad Hoc")
        ws.cell(row=row_idx, column=7, value="")

    # Auto-width columns
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

    wb.save(output_path)
    progress.log(f"Excel written to {output_path}")


# ---------------------------------------------------------------------------
# Main orchestrator
# ---------------------------------------------------------------------------

def scan(website_url: str, output_dir: str = ".", progress_callback=None,
         mode: str = "homepage", interaction: InteractionRequest = None,
         jurisdiction_name: str = "") -> dict:
    """
    Full pipeline: validate → discover → enumerate → filter → deduplicate → export.

    Args:
        website_url: Root URL of the government website (or REST directory).
        output_dir: Directory for output files.
        progress_callback: Optional callable(event_type, message) for live updates.
        mode: Workflow mode selected by user:
            "direct"   – URL is an ArcGIS REST Services Directory root (Path A).
            "homepage" – URL is the jurisdiction's website (Path B: LLM discovery).
        interaction: Optional InteractionRequest for mid-scan user prompts.
        jurisdiction_name: Optional human-readable name (e.g., "City of Las Vegas").
            If empty, derived from the URL domain.

    Returns:
        dict with keys: xl_path, md_path, stats, error (if any).
    """
    progress.reset(callback=progress_callback)

    # Re-read the planning layer pattern skill document so any keyword
    # updates committed since the server started are picked up immediately.
    _load_keywords_from_doc()

    print("\n" + "=" * 60)
    print("  Government ArcGIS Feature Layer Scanner")
    print("=" * 60 + "\n")

    # Step 0 – Validate URL (skip if user provided only a jurisdiction name)
    has_url = website_url and website_url.startswith(("http://", "https://"))
    if has_url:
        progress.log(f"Validating URL: {website_url}")
        validation = validate_url(website_url, mode=mode)
        if validation["valid"]:
            progress.log(f"✓ URL is valid and reachable — {validation['message']}")
        else:
            progress.log(f"✗ URL validation failed — {validation['message']}")
            if progress_callback:
                progress_callback("error_msg", f"URL validation failed: {validation['message']}")
            progress.summary()
            return {"error": f"URL validation failed: {validation['message']}",
                    "stats": dict(progress.stats)}
    elif jurisdiction_name:
        progress.log(f"Searching for: {jurisdiction_name}")
    else:
        progress.summary()
        return {"error": "No URL or jurisdiction name provided.",
                "stats": dict(progress.stats)}

    # Step 1 – find ArcGIS REST endpoints (skipped for "direct" mode)
    if mode == "direct":
        # Path A: user provided the REST Services Directory URL directly
        rest_url = website_url.rstrip("/")
        # Normalise to the root services directory
        idx = rest_url.lower().find("/rest/services")
        if idx != -1:
            rest_url = rest_url[: idx + len("/rest/services")]
        rest_urls = {rest_url}
        progress.log(f"Direct mode: using ArcGIS REST directory at {rest_url}")
    else:
        # Path B: search for ArcGIS REST endpoints via LLM web search
        if not jurisdiction_name:
            # Derive a human-readable name from the domain
            parsed_url = urlparse(website_url)
            domain = parsed_url.netloc.replace("www.", "")
            # e.g., "cityoflasvegas.com" → "cityoflasvegas" → "city of las vegas"
            #        "milpitas.gov" → "milpitas" → "City of Milpitas"
            #        "co.surry.nc.us" → "co surry nc" → "Surry County"
            # Strip TLD and common suffixes to get the jurisdiction slug
            # e.g., "milpitas.gov" → "milpitas", "co.surry.nc.us" → "co.surry.nc"
            #        "cityoflasvegas.com" → "cityoflasvegas"
            for suffix in [".gov", ".us", ".org", ".com", ".net"]:
                if domain.endswith(suffix):
                    domain = domain[: -len(suffix)]
                    break
            # Remove state codes (2-letter segments) from multi-part domains
            # e.g., "sunnyvale.ca" → "sunnyvale", "co.surry.nc" → "co.surry"
            _US_STATES = {
                "al", "ak", "az", "ar", "ca", "co", "ct", "de", "fl", "ga",
                "hi", "id", "il", "in", "ia", "ks", "ky", "la", "me", "md",
                "ma", "mi", "mn", "ms", "mo", "mt", "ne", "nv", "nh", "nj",
                "nm", "ny", "nc", "nd", "oh", "ok", "or", "pa", "ri", "sc",
                "sd", "tn", "tx", "ut", "vt", "va", "wa", "wv", "wi", "wy",
            }
            # Keep "co" (county prefix) but remove state codes
            parts = [p for p in domain.split(".")
                     if p.lower() not in _US_STATES or p.lower() == "co"]
            jurisdiction_slug = " ".join(parts)
            jurisdiction_slug = jurisdiction_slug.replace("-", " ").replace("_", " ")
            # Insert spaces before capital letters (e.g., "cityoflasvegas" → "cityof las vegas")
            jurisdiction_name = re.sub(r"([a-z])([A-Z])", r"\1 \2", jurisdiction_slug)
            # Expand common prefixes
            jurisdiction_name = re.sub(r"\bco\b", "County of", jurisdiction_name, flags=re.IGNORECASE)
            jurisdiction_name = re.sub(r"\bcityof\b", "City of", jurisdiction_name, flags=re.IGNORECASE)
            # If it's a single word (like "milpitas"), prepend "City of"
            words = jurisdiction_name.strip().split()
            if len(words) == 1 and not any(w in jurisdiction_name.lower() for w in ["county", "city", "town"]):
                jurisdiction_name = f"City of {jurisdiction_name}"
            jurisdiction_name = jurisdiction_name.strip().title()

        # SECURITY: Sanitise jurisdiction name before passing to LLM
        jurisdiction_name = sanitize_jurisdiction_name(jurisdiction_name)

        progress.log(f"Discovery mode: finding ArcGIS endpoints for {jurisdiction_name}")
        rest_urls = llm_search_for_arcgis(
            jurisdiction_name=jurisdiction_name,
            homepage_url=website_url if has_url else "",
        )

    if not rest_urls:
        progress.log("ERROR: Could not discover any ArcGIS REST Services Directory.")
        progress.summary()
        return {"error": "No ArcGIS REST endpoints found.", "stats": dict(progress.stats)}

    # Step 2 – enumerate all layers across discovered endpoints
    all_layers: list[dict] = []
    for rest_url in rest_urls:
        all_layers.extend(query_rest_services(rest_url))

    if not all_layers:
        progress.log("ERROR: No feature layers found at the discovered endpoints.")
        progress.summary()
        return {"error": "No feature layers found.", "stats": dict(progress.stats)}

    # Step 3 – filter
    filtered = filter_layers(all_layers)
    if not filtered:
        progress.log("WARNING: No layers matched planning/development keywords. "
                      "Outputting all non-excluded layers instead.")
        filtered = [l for l in all_layers if not is_excluded_by_pattern(l["layer_name"])
                    and not is_hard_excluded(l["layer_name"])
                    and not is_service_excluded(l["service_name"])]
        for l in filtered:
            l["priority"] = 99

    # Step 4 – deduplicate
    final_layers = deduplicate(filtered)

    # Step 4b – fetch layer details (record count, geometry) only for final set
    enrich_layer_details(final_layers)

    # Step 5 – output
    os.makedirs(output_dir, exist_ok=True)

    if has_url:
        file_slug = urlparse(website_url).netloc.replace("www.", "").split(".")[0]
    else:
        file_slug = re.sub(r"[^a-z0-9]+", "_", jurisdiction_name.lower()).strip("_")
    md_path = os.path.join(output_dir, f"{file_slug}_feature_layers.md")
    xl_path = os.path.join(output_dir, f"{file_slug}_feature_layers.xlsx")

    write_markdown(final_layers, md_path)
    write_excel(final_layers, xl_path)

    progress.stat("Final layers exported", len(final_layers))
    progress.summary()

    print(f"  Output files:")
    print(f"    Markdown : {os.path.abspath(md_path)}")
    print(f"    Excel    : {os.path.abspath(xl_path)}")
    print()

    return {
        "xl_path": os.path.abspath(xl_path),
        "md_path": os.path.abspath(md_path),
        "stats": dict(progress.stats),
        "layers": final_layers,
    }


# ---------------------------------------------------------------------------
# CLI entry point
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Scan a local government website for ArcGIS planning/development feature layers."
    )
    parser.add_argument(
        "input",
        help="ArcGIS REST URL, jurisdiction website URL, or jurisdiction name "
             "(e.g. 'https://gis.example.gov/arcgis/rest/services' or 'City of Sunnyvale')",
    )
    parser.add_argument(
        "-o", "--output-dir",
        default=".",
        help="Directory for output files (default: current directory)",
    )
    parser.add_argument(
        "-m", "--mode",
        choices=["direct", "homepage"],
        default=None,
        help="Scan mode: 'direct' for ArcGIS REST URL, 'homepage' for discovery. "
             "Auto-detected from input if not specified.",
    )
    args = parser.parse_args()

    user_input = args.input
    is_url = user_input.startswith(("http://", "https://"))

    # Auto-detect mode
    mode = args.mode
    if mode is None:
        if is_url and "/rest/services" in user_input.lower():
            mode = "direct"
        else:
            mode = "homepage"

    # Determine URL vs jurisdiction name
    url = user_input if is_url else ""
    jurisdiction_name = "" if is_url else user_input

    result = scan(url, args.output_dir, mode=mode,
                  jurisdiction_name=jurisdiction_name)
    if result.get("error"):
        sys.exit(1)


if __name__ == "__main__":
    main()
